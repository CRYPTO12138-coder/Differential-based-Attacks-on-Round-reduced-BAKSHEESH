from gurobipy import *
import openpyxl
from openpyxl.styles import PatternFill, Alignment

# --- 1. Visualize X Variables (Rounds 1-5) ---
def SaveXToExcel(model, excelFileName):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "X Visualization"
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data = {}
    for v in model.getVars():
        if v.VarName.startswith("X_"):
            parts = v.VarName.split('_')
            r, bit = int(parts[1]), int(parts[2])
            data[(r, bit)] = int(round(v.X))

    sorted_rounds = sorted(list(set(k[0] for k in data.keys())))
    for r_idx, r_val in enumerate(sorted_rounds):
        ws.cell(row=r_idx + 1, column=1).value = f"X R{r_val}"
        for b_idx in range(128):
            val = data.get((r_val, b_idx), 0)
            col_idx = 128 - b_idx + 1
            cell = ws.cell(row=r_idx + 1, column=col_idx)
            cell.value = val
            cell.fill = red_fill if val == 1 else white_fill
            cell.alignment = Alignment(horizontal='center')

    for col in range(2, 130): ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 2.5
    wb.save(excelFileName)
    print(f"X matrix generated: {excelFileName}")

# --- 1. Visualize Y Variables (Rounds 1-5) ---
def SaveYToExcel(model, excelFileName):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Y Visualization"
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data = {}
    for v in model.getVars():
        if v.VarName.startswith("Y_"):
            parts = v.VarName.split('_')
            r, bit = int(parts[1]), int(parts[2])
            data[(r, bit)] = int(round(v.X))

    sorted_rounds = sorted(list(set(k[0] for k in data.keys())))
    for r_idx, r_val in enumerate(sorted_rounds):
        ws.cell(row=r_idx + 1, column=1).value = f"Y R{r_val}"
        for b_idx in range(128):
            val = data.get((r_val, b_idx), 0)
            col_idx = 128 - b_idx + 1
            cell = ws.cell(row=r_idx + 1, column=col_idx)
            cell.value = val
            cell.fill = red_fill if val == 1 else white_fill
            cell.alignment = Alignment(horizontal='center')

    for col in range(2, 130): ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 2.5
    wb.save(excelFileName)
    print(f"Y matrix generated: {excelFileName}")

# --- 2. Visualize BX0 and BX1 Variables (Backward Rounds) ---
def SaveBXToExcel(model, excelFileName):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BX Visualization"
    color_B = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # (0,1) blue
    color_C = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # (1,0) red
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    bx_data = {}
    for v in model.getVars():
        if v.VarName.startswith("BX0_") or v.VarName.startswith("BX1_"):
            parts = v.VarName.split('_')
            v_type, r, bit = parts[0], int(parts[1]), int(parts[2])
            key = (r, bit)
            if key not in bx_data: bx_data[key] = [0, 0]
            bx_data[key][0 if v_type == "BX0" else 1] = int(round(v.X))

    sorted_rounds = sorted(list(set(k[0] for k in bx_data.keys())))
    for r_idx, r_val in enumerate(sorted_rounds):
        ws.cell(row=r_idx + 1, column=1).value = f"BX R{r_val}"
        for b_idx in range(128):
            bx0, bx1 = bx_data.get((r_val, b_idx), [0, 0])
            col_idx = 128 - b_idx + 1
            cell = ws.cell(row=r_idx + 1, column=col_idx)
            cell.value = f"{bx0}{bx1}"
            cell.alignment = Alignment(horizontal='center')
            if bx0 == 0 and bx1 == 1: cell.fill = color_B
            elif bx0 == 1 and bx1 == 0: cell.fill = color_C
            else: cell.fill = white_fill

    for col in range(2, 130): ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 3.5
    wb.save(excelFileName)
    print(f"BX matrix generated: {excelFileName}")

# --- 2. Visualize BY0 and BY1 Variables (Backward Rounds) ---
def SaveBYToExcel(model, excelFileName):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BY Visualization"
    color_B = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid") # (0,1) 蓝色
    color_C = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # (1,0) 红色
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    bx_data = {}
    for v in model.getVars():
        if v.VarName.startswith("BY0_") or v.VarName.startswith("BY1_"):
            parts = v.VarName.split('_')
            v_type, r, bit = parts[0], int(parts[1]), int(parts[2])
            key = (r, bit)
            if key not in bx_data: bx_data[key] = [0, 0]
            bx_data[key][0 if v_type == "BY0" else 1] = int(round(v.X))

    sorted_rounds = sorted(list(set(k[0] for k in bx_data.keys())))
    for r_idx, r_val in enumerate(sorted_rounds):
        ws.cell(row=r_idx + 1, column=1).value = f"BX R{r_val}"
        for b_idx in range(128):
            bx0, bx1 = bx_data.get((r_val, b_idx), [0, 0])
            col_idx = 128 - b_idx + 1
            cell = ws.cell(row=r_idx + 1, column=col_idx)
            cell.value = f"{bx0}{bx1}"
            cell.alignment = Alignment(horizontal='center')
            if bx0 == 0 and bx1 == 1: cell.fill = color_B
            elif bx0 == 1 and bx1 == 0: cell.fill = color_C
            else: cell.fill = white_fill

    for col in range(2, 130): ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 3.5
    wb.save(excelFileName)
    print(f"BY matrix generated: {excelFileName}")

# --- 3. Visualize XX Variables ---
def SaveXXToExcel(model, excelFileName):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "XX Visualization"
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    data = {}
    for v in model.getVars():
        if v.VarName.startswith("XX_"):
            parts = v.VarName.split('_')
            r, bit = int(parts[1]), int(parts[2])
            data[(r, bit)] = int(round(v.X))

    sorted_rounds = sorted(list(set(k[0] for k in data.keys())))
    for r_idx, r_val in enumerate(sorted_rounds):
        ws.cell(row=r_idx + 1, column=1).value = f"XX R{r_val}"
        max_bit = max([k[1] for k in data.keys() if k[0] == r_val])
        for b_idx in range(max_bit + 1):
            val = data.get((r_val, b_idx), 0)
            col_idx = (max_bit + 1) - b_idx + 1
            cell = ws.cell(row=r_idx + 1, column=col_idx)
            cell.value = val
            cell.fill = yellow_fill if val == 1 else white_fill
            cell.alignment = Alignment(horizontal='center')

    wb.save(excelFileName)
    print(f"XX matrix generated: {excelFileName}")

# --- 4. Visualize KK Variables ---
def SaveKKToExcel(model, excelFileName):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KK Visualization"
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # 1 为黄色
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    FIXED_STATE_LENGTH = 128

    data = {}
    for v in model.getVars():
        if v.VarName.startswith("KK_"):
            try:
                parts = v.VarName.split('_')
                r = int(parts[1])
                bit = int(parts[2])
                val = int(round(v.X))
                data[(r, bit)] = val
            except:
                continue

    if not data:
        print("No KK_ variables found in the model.")
        return

    sorted_rounds = sorted(list(set(k[0] for k in data.keys())))

    for r_idx, r_val in enumerate(sorted_rounds):
        ws.cell(row=r_idx + 1, column=1).value = f"KK Round {r_val}"
        
        for b_idx in range(FIXED_STATE_LENGTH):
            val = data.get((r_val, b_idx), 0)
            
            col_idx = (FIXED_STATE_LENGTH - 1 - b_idx) + 2
            
            cell = ws.cell(row=r_idx + 1, column=col_idx)
            cell.value = val
            cell.alignment = Alignment(horizontal='center')
            
            if val == 1:
                cell.fill = yellow_fill
            else:
                cell.fill = white_fill

    ws.column_dimensions['A'].width = 15
    for col in range(2, FIXED_STATE_LENGTH + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 2.5

    wb.save(excelFileName)
    print(f"KK visualization saved to: {excelFileName}")

##################################################MODEL##################################################
def GetVariables(round,varName,varSize,variable):
    res = []
    for i in range(varSize):
        res.append(varName + "_" + str(round) + "_" + str(i))
        variable.add(varName + "_" + str(round) + "_" + str(i))
    return res

def Constraint_initialize_FD0(f, variable):
    f.write("c" + " = 1 " + "\n")

    res = []
    for i in range (STATE_LENGTH):
        res.append(GetVariables(5,"X",STATE_LENGTH,variable)[i])
        f.write(GetVariables(5,"X",STATE_LENGTH,variable)[i] + " - " + str((AA[aa][0]>>i)&0x1) + " c = 0\n")
    f.write(" + ".join(res) +  " - " + str(AA[aa][2]) + " c = 0 " + "\n")

    res = []
    for r in range (1, 5):
        for i in range (int(STATE_LENGTH/4)):
            res.append("2 " + GetVariables(r,"p_D",int(STATE_LENGTH/4),variable)[i])
    for r in range (11, 16):
        for i in range (int(STATE_LENGTH/4)):
            res.append("2 " + GetVariables(r,"p_D",int(STATE_LENGTH/4),variable)[i])
    f.write(" + ".join(res) + " <= 36\n")

    for r in range (0, 2):
        for i in range (32):
            f.write(GetVariables(r,"XX",32,variable)[i] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+0] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+0] + " >= 0\n")
            f.write(GetVariables(r,"XX",32,variable)[i] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+1] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+1] + " >= 0\n")
            f.write(GetVariables(r,"XX",32,variable)[i] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+2] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+2] + " >= 0\n")
            f.write(GetVariables(r,"XX",32,variable)[i] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+3] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+3] + " >= 0\n")
            f.write(GetVariables(r,"XX",32,variable)[i] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+0] \
            + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+1] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+2] \
            + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+3] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+0] \
            + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+1] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+2] \
            + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+3] + " <= 0\n")

    for i in range (32):
        f.write(GetVariables(0,"XX",32,variable)[i] + " - " + GetVariables(0,"KK",STATE_LENGTH,variable)[(4*i+0)%STATE_LENGTH] + " = 0\n")
        f.write(GetVariables(0,"XX",32,variable)[i] + " - " + GetVariables(0,"KK",STATE_LENGTH,variable)[(4*i+1)%STATE_LENGTH] + " = 0\n")
        f.write(GetVariables(0,"XX",32,variable)[i] + " - " + GetVariables(0,"KK",STATE_LENGTH,variable)[(4*i+2)%STATE_LENGTH] + " = 0\n")
 
####################################################################################################
def Constraint_sbox_FD0(r, f, variable):
    P = [0, 33, 66, 99, 96, 1, 34, 67, 64, 97, 2, 35, 32, 65, 98, 3, 4, 37, 70, 103, 100, 5, 38, 71, 68, 101, 6, 39, 36, 69, 102, 7, 8, 41, 
    74, 107, 104, 9, 42, 75, 72, 105, 10, 43, 40, 73, 106, 11, 12, 45, 78, 111, 108, 13, 46, 79, 76, 109, 14, 47, 44, 77, 110, 15, 16, 49, 
    82, 115, 112, 17, 50, 83, 80, 113, 18, 51, 48, 81, 114, 19, 20, 53, 86, 119, 116, 21, 54, 87, 84, 117, 22, 55, 52, 85, 118, 23, 24, 57, 
    90, 123, 120, 25, 58, 91, 88, 121, 26, 59, 56, 89, 122, 27, 28, 61, 94, 127, 124, 29, 62, 95, 92, 125, 30, 63, 60, 93, 126, 31]

    M0 = [[2, 1, -3, -3, -2, 2, -1, -1, 7, 0], [-3, -2, 0, -2, 2, 1, 1, -1, 6, 0], [-1, 0, 0, 0, -2, -1, -1, -1, -2, -6], [1, 0, 0, 0, 2, 1, 1, 1, -2, 0], [-1, 2, 0, 0, 4, -1, -1, -1, 2, 0], [1, 2, 0, 0, -4, 1, 1, 1, 2, 0], [0, 1, 1, 1, 0, 0, 0, 0, -1, 0], [1, 1, 0, 3, 2, -2, -2, 1, 1, 0], [2, 1, 3, 0, -2, -1, 2, -1, 1, 0], [-1, 1, 0, 3, -2, 2, 2, -1, 1, 0], [-2, 1, 3, 0, 2, 1, -2, 1, 1, 0], [-1, -2, -2, 0, -2, 1, -1, 1, 4, -2], [1, -2, 0, -2, -2, -1, -1, 1, 4, -2], [1, -2, -2, 0, 2, -3, 1, -1, 6, 0], [-2, 1, -3, -3, 2, -2, 1, 1, 7, 0], [0, -1, -1, 1, 1, 1, -1, -1, 3, 0], [1, -1, 1, -1, 1, 0, -1, -1, 3, 0], [-1, 0, 1, 1, 1, -1, 1, 0, 1, 0], [1, 0, 1, 1, -1, 1, -1, 0, 1, 0], [-1, -1, 1, -1, -1, 0, 1, 1, 3, 0], [0, -1, -1, 1, -1, -1, 1, 1, 3, 0]]

    for i in range (int(STATE_LENGTH/4)):
        for t in range (len(M0)):
            res = []
            res.append(str(M0[t][0]) + " " + GetVariables(r,"X",STATE_LENGTH,variable)[4*i+3])
            res.append(str(M0[t][1]) + " " + GetVariables(r,"X",STATE_LENGTH,variable)[4*i+2])
            res.append(str(M0[t][2]) + " " + GetVariables(r,"X",STATE_LENGTH,variable)[4*i+1])
            res.append(str(M0[t][3]) + " " + GetVariables(r,"X",STATE_LENGTH,variable)[4*i+0])
            res.append(str(M0[t][4]) + " " + GetVariables(r+1,"X",STATE_LENGTH,variable)[P[4*i+3]])
            res.append(str(M0[t][5]) + " " + GetVariables(r+1,"X",STATE_LENGTH,variable)[P[4*i+2]])
            res.append(str(M0[t][6]) + " " + GetVariables(r+1,"X",STATE_LENGTH,variable)[P[4*i+1]])
            res.append(str(M0[t][7]) + " " + GetVariables(r+1,"X",STATE_LENGTH,variable)[P[4*i+0]])
            res.append(str(M0[t][8]) + " " + GetVariables(r,"p_D",int(STATE_LENGTH/4),variable)[i])
            f.write(" + ".join(res) + " - " + str(M0[t][9]) + " c" + " >= 0 " + "\n")
    for i in range (STATE_LENGTH):
        f.write(GetVariables(r,"Y",STATE_LENGTH,variable)[i] + " - " + GetVariables(r+1,"X",STATE_LENGTH,variable)[P[i]] + " = 0\n")

def Constraint_FD0(f, variable):
    Constraint_initialize_FD0(f, variable)
    for r in range (1, 5):
        Constraint_sbox_FD0(r, f, variable)


def Constraint_initialize_MBD0(f, variable):
    f.write("c" + " = 1 " + "\n")

    res = []
    for i in range (STATE_LENGTH):
        res.append(GetVariables(1,"BX0",STATE_LENGTH,variable)[i])
        res.append(GetVariables(1,"BX1",STATE_LENGTH,variable)[i])
    f.write(" + ".join(res) + " >= 1 " + "\n")

    for i in range (STATE_LENGTH):
        f.write(GetVariables(1,"BX0",STATE_LENGTH,variable)[i] + " = 0\n")
        f.write(GetVariables(1,"BX1",STATE_LENGTH,variable)[i] + " - " + GetVariables(1,"X",STATE_LENGTH,variable)[i] + " = 0\n")

def Constraint_sbox_MBD0(r, f, variable):
    P = [0, 33, 66, 99, 96, 1, 34, 67, 64, 97, 2, 35, 32, 65, 98, 3, 4, 37, 70, 103, 100, 5, 38, 71, 68, 101, 6, 39, 36, 69, 102, 7, 8, 41, 
    74, 107, 104, 9, 42, 75, 72, 105, 10, 43, 40, 73, 106, 11, 12, 45, 78, 111, 108, 13, 46, 79, 76, 109, 14, 47, 44, 77, 110, 15, 16, 49, 
    82, 115, 112, 17, 50, 83, 80, 113, 18, 51, 48, 81, 114, 19, 20, 53, 86, 119, 116, 21, 54, 87, 84, 117, 22, 55, 52, 85, 118, 23, 24, 57, 
    90, 123, 120, 25, 58, 91, 88, 121, 26, 59, 56, 89, 122, 27, 28, 61, 94, 127, 124, 29, 62, 95, 92, 125, 30, 63, 60, 93, 126, 31]

    M0 = [[-1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, -1, 0], 
    [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, 0, 0, -1], [0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0], 
    [0, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, -1], 
    [0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0], [0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1], 
    [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, -1, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, -1], 
    [0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, -1], [0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0], 
    [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, -1, 0, 0, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1], 
    [-1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0], [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, -1, 0, 0, 0], 
    [0, 0, 0, 0, 0, 0, 0, 1, 0, -1, 0, 0, 0, 0, 1, 0, 0], [0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 1, 0, 0], 
    [0, 1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 0], [-1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1], 
    [0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, -2], [0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0], 
    [0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1], [0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 0, -1, 0, 0, 0, 0, -1], 
    [0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1], [0, 0, 1, 0, 1, 0, 1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0], 
    [1, 0, 0, 0, 0, -1, 0, -1, 0, -1, -1, 0, 0, 0, 0, 0, -3], [0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 0, 0, 1, 0, 0, 0, 0], 
    [0, 0, 0, 0, 0, 0, 0, -1, 0, 1, 0, -1, 0, 0, 0, 0, -1], [1, 0, 0, 1, 0, 1, 0, 0, 1, 1, -1, 0, 0, 0, 0, 0, 0], 
    [0, 1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, -1], [0, 0, 0, 1, 0, -1, 0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0], 
    [0, -1, 0, 0, 0, 1, 0, -1, 1, 0, 0, 0, 0, 1, 0, 1, -1], [0, 0, 0, -1, 0, 1, 0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0], 
    [0, -1, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], [0, 0, 0, 1, 0, 1, 0, 1, 0, -1, 0, 0, 0, 0, 0, 0, 0], 
    [0, 0, 0, -1, 0, 1, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, -2], [1, 1, 0, 0, 0, -1, 0, 1, 1, 0, 0, 0, 0, 1, 0, 1, 0], 
    [0, 0, 0, -1, 0, -1, 0, -1, 0, 1, 0, 0, 0, 0, 0, 0, -2], [-1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1], 
    [0, 0, 0, -1, 0, -1, 0, 1, 0, -1, 0, 0, 0, 0, 0, 0, -2], [0, 0, 0, 1, 0, -1, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, -2], 
    [0, 0, 0, 1, 0, 1, 0, -1, 1, 1, 0, 0, 0, 0, 0, 0, 0]]
    for i in range (int(STATE_LENGTH/4)):
        for t in range (len(M0)):
            res = []
            res.append(str(M0[t][0]) + " " + GetVariables(r+1,"BX0",STATE_LENGTH,variable)[P[4*i+3]])
            res.append(str(M0[t][1]) + " " + GetVariables(r+1,"BX1",STATE_LENGTH,variable)[P[4*i+3]])
            res.append(str(M0[t][2]) + " " + GetVariables(r+1,"BX0",STATE_LENGTH,variable)[P[4*i+2]])
            res.append(str(M0[t][3]) + " " + GetVariables(r+1,"BX1",STATE_LENGTH,variable)[P[4*i+2]])
            res.append(str(M0[t][4]) + " " + GetVariables(r+1,"BX0",STATE_LENGTH,variable)[P[4*i+1]])
            res.append(str(M0[t][5]) + " " + GetVariables(r+1,"BX1",STATE_LENGTH,variable)[P[4*i+1]])
            res.append(str(M0[t][6]) + " " + GetVariables(r+1,"BX0",STATE_LENGTH,variable)[P[4*i+0]])
            res.append(str(M0[t][7]) + " " + GetVariables(r+1,"BX1",STATE_LENGTH,variable)[P[4*i+0]])
            res.append(str(M0[t][8]) + " " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+3])
            res.append(str(M0[t][9]) + " " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+3])
            res.append(str(M0[t][10]) + " " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+2])
            res.append(str(M0[t][11]) + " " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+2])
            res.append(str(M0[t][12]) + " " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+1])
            res.append(str(M0[t][13]) + " " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+1])
            res.append(str(M0[t][14]) + " " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+0])
            res.append(str(M0[t][15]) + " " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+0])
            f.write(" + ".join(res) + " - " + str(M0[t][16]) + " c" + " >= 0 " + "\n")

    for i in range (STATE_LENGTH):
        f.write(GetVariables(r,"BY0",STATE_LENGTH,variable)[i] + " - " + GetVariables(r+1,"BX0",STATE_LENGTH,variable)[P[i]] + " = 0\n")
        f.write(GetVariables(r,"BY1",STATE_LENGTH,variable)[i] + " - " + GetVariables(r+1,"BX1",STATE_LENGTH,variable)[P[i]] + " = 0\n")

def Constraint_M_BD0(f, variable):
    Constraint_initialize_MBD0(f, variable)
    for r in range (0, 1):
        Constraint_sbox_MBD0(r, f, variable)


def Constraint_initialize_FD1(f, variable):
    f.write("c" + " = 1 " + "\n")

    res = []
    for i in range (STATE_LENGTH):
        res.append(GetVariables(11,"X",STATE_LENGTH,variable)[i])
        f.write(GetVariables(11,"X",STATE_LENGTH,variable)[i] + " - " + str((AA[aa][1]>>i)&0x1) + " c = 0\n")
    f.write(" + ".join(res) +  " - " + str(AA[aa][3]) + " c = 0 " + "\n")

    for r in range (16, 19):
        for i in range (32):
            f.write(GetVariables(r,"XX",32,variable)[i] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+0] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+0] + " >= 0\n")
            f.write(GetVariables(r,"XX",32,variable)[i] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+1] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+1] + " >= 0\n")
            f.write(GetVariables(r,"XX",32,variable)[i] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+2] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+2] + " >= 0\n")
            f.write(GetVariables(r,"XX",32,variable)[i] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+3] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+3] + " >= 0\n")
            f.write(GetVariables(r,"XX",32,variable)[i] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+0] \
            + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+1] + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+2] \
            + " - " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+3] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+0] \
            + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+1] + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+2] \
            + " - " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+3] + " <= 0\n")
    P = [0, 33, 66, 99, 96, 1, 34, 67, 64, 97, 2, 35, 32, 65, 98, 3, 4, 37, 70, 103, 100, 5, 38, 71, 68, 101, 6, 39, 36, 69, 102, 7, 8, 41, 
    74, 107, 104, 9, 42, 75, 72, 105, 10, 43, 40, 73, 106, 11, 12, 45, 78, 111, 108, 13, 46, 79, 76, 109, 14, 47, 44, 77, 110, 15, 16, 49, 
    82, 115, 112, 17, 50, 83, 80, 113, 18, 51, 48, 81, 114, 19, 20, 53, 86, 119, 116, 21, 54, 87, 84, 117, 22, 55, 52, 85, 118, 23, 24, 57, 
    90, 123, 120, 25, 58, 91, 88, 121, 26, 59, 56, 89, 122, 27, 28, 61, 94, 127, 124, 29, 62, 95, 92, 125, 30, 63, 60, 93, 126, 31]
    M0 = [[0, -1, 0, 0, 0, 0, 0, 1, 0], [0, 0, 0, 0, 1, 0, 0, -1, 0], [0, 0, 0, 0, -1, 0, 1, 0, 0], [0, 0, 0, 0, 0, 1, -1, 0, 0], [0, 0, -1, 0, 0, 0, 0, 1, 0], [0, 1, 1, 1, 0, -1, 0, 0, 0], [0, 0, 0, -1, 0, 0, 0, 1, 0]]
    for i in range (int(STATE_LENGTH/4)):
        for t in range (len(M0)):
            res = []
            res.append(str(M0[t][0]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[4*i+3])
            res.append(str(M0[t][1]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[4*i+2])
            res.append(str(M0[t][2]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[4*i+1])
            res.append(str(M0[t][3]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[4*i+0])
            res.append(str(M0[t][4]) + " " + GetVariables(17,"KK",STATE_LENGTH,variable)[(P[4*i+3]+17)%128])
            res.append(str(M0[t][5]) + " " + GetVariables(17,"KK",STATE_LENGTH,variable)[(P[4*i+2]+17)%128])
            res.append(str(M0[t][6]) + " " + GetVariables(17,"KK",STATE_LENGTH,variable)[(P[4*i+1]+17)%128])
            res.append(str(M0[t][7]) + " " + GetVariables(17,"KK",STATE_LENGTH,variable)[(P[4*i+0]+17)%128])
            f.write(" + ".join(res) + " - " + str(M0[t][8]) + " c" + " >= 0 " + "\n")

    for i in range (STATE_LENGTH):
        f.write(GetVariables(19,"KK",STATE_LENGTH,variable)[i] + " - " + GetVariables(18,"KK",STATE_LENGTH,variable)[i] + " >= 0 " + "\n")
        f.write(GetVariables(19,"KK",STATE_LENGTH,variable)[i] + " - " + GetVariables(17,"KK",STATE_LENGTH,variable)[i] + " >= 0 " + "\n")
        f.write(GetVariables(19,"KK",STATE_LENGTH,variable)[i] + " - " + GetVariables(0,"KK",STATE_LENGTH,variable)[i] + " >= 0 " + "\n")
        f.write(GetVariables(19,"KK",STATE_LENGTH,variable)[i] + " - " + GetVariables(18,"KK",STATE_LENGTH,variable)[i] + " - " + GetVariables(17,"KK",STATE_LENGTH,variable)[i] + " - " + GetVariables(0,"KK",STATE_LENGTH,variable)[i] + " <= 0 " + "\n")

    for i in range (32):
        f.write(GetVariables(17,"XX",32,variable)[i] + " - " + GetVariables(18,"KK",STATE_LENGTH,variable)[(P[4*i+0]+18)%STATE_LENGTH] + " = 0\n")
        f.write(GetVariables(17,"XX",32,variable)[i] + " - " + GetVariables(18,"KK",STATE_LENGTH,variable)[(P[4*i+1]+18)%STATE_LENGTH] + " = 0\n")
        f.write(GetVariables(17,"XX",32,variable)[i] + " - " + GetVariables(18,"KK",STATE_LENGTH,variable)[(P[4*i+2]+18)%STATE_LENGTH] + " = 0\n")
        f.write(GetVariables(17,"XX",32,variable)[i] + " - " + GetVariables(18,"KK",STATE_LENGTH,variable)[(P[4*i+3]+18)%STATE_LENGTH] + " = 0\n")

####################################################################################################
def Constraint_sbox_FD1(r, f, variable):
    P = [0, 33, 66, 99, 96, 1, 34, 67, 64, 97, 2, 35, 32, 65, 98, 3, 4, 37, 70, 103, 100, 5, 38, 71, 68, 101, 6, 39, 36, 69, 102, 7, 8, 41, 
    74, 107, 104, 9, 42, 75, 72, 105, 10, 43, 40, 73, 106, 11, 12, 45, 78, 111, 108, 13, 46, 79, 76, 109, 14, 47, 44, 77, 110, 15, 16, 49, 
    82, 115, 112, 17, 50, 83, 80, 113, 18, 51, 48, 81, 114, 19, 20, 53, 86, 119, 116, 21, 54, 87, 84, 117, 22, 55, 52, 85, 118, 23, 24, 57, 
    90, 123, 120, 25, 58, 91, 88, 121, 26, 59, 56, 89, 122, 27, 28, 61, 94, 127, 124, 29, 62, 95, 92, 125, 30, 63, 60, 93, 126, 31]

    M0 = [[2, 1, -3, -3, -2, 2, -1, -1, 7, 0], [-3, -2, 0, -2, 2, 1, 1, -1, 6, 0], [-1, 0, 0, 0, -2, -1, -1, -1, -2, -6], [1, 0, 0, 0, 2, 1, 1, 1, -2, 0], [-1, 2, 0, 0, 4, -1, -1, -1, 2, 0], [1, 2, 0, 0, -4, 1, 1, 1, 2, 0], [0, 1, 1, 1, 0, 0, 0, 0, -1, 0], [1, 1, 0, 3, 2, -2, -2, 1, 1, 0], [2, 1, 3, 0, -2, -1, 2, -1, 1, 0], [-1, 1, 0, 3, -2, 2, 2, -1, 1, 0], [-2, 1, 3, 0, 2, 1, -2, 1, 1, 0], [-1, -2, -2, 0, -2, 1, -1, 1, 4, -2], [1, -2, 0, -2, -2, -1, -1, 1, 4, -2], [1, -2, -2, 0, 2, -3, 1, -1, 6, 0], [-2, 1, -3, -3, 2, -2, 1, 1, 7, 0], [0, -1, -1, 1, 1, 1, -1, -1, 3, 0], [1, -1, 1, -1, 1, 0, -1, -1, 3, 0], [-1, 0, 1, 1, 1, -1, 1, 0, 1, 0], [1, 0, 1, 1, -1, 1, -1, 0, 1, 0], [-1, -1, 1, -1, -1, 0, 1, 1, 3, 0], [0, -1, -1, 1, -1, -1, 1, 1, 3, 0]]

    for i in range (int(STATE_LENGTH/4)):
        for t in range (len(M0)):
            res = []
            res.append(str(M0[t][0]) + " " + GetVariables(r,"X",STATE_LENGTH,variable)[4*i+3])
            res.append(str(M0[t][1]) + " " + GetVariables(r,"X",STATE_LENGTH,variable)[4*i+2])
            res.append(str(M0[t][2]) + " " + GetVariables(r,"X",STATE_LENGTH,variable)[4*i+1])
            res.append(str(M0[t][3]) + " " + GetVariables(r,"X",STATE_LENGTH,variable)[4*i+0])
            res.append(str(M0[t][4]) + " " + GetVariables(r+1,"X",STATE_LENGTH,variable)[P[4*i+3]])
            res.append(str(M0[t][5]) + " " + GetVariables(r+1,"X",STATE_LENGTH,variable)[P[4*i+2]])
            res.append(str(M0[t][6]) + " " + GetVariables(r+1,"X",STATE_LENGTH,variable)[P[4*i+1]])
            res.append(str(M0[t][7]) + " " + GetVariables(r+1,"X",STATE_LENGTH,variable)[P[4*i+0]])
            res.append(str(M0[t][8]) + " " + GetVariables(r,"p_D",int(STATE_LENGTH/4),variable)[i])
            f.write(" + ".join(res) + " - " + str(M0[t][9]) + " c" + " >= 0 " + "\n")

def Constraint_FD1(f, variable):
    Constraint_initialize_FD1(f, variable)
    for r in range (11, 17):
        Constraint_sbox_FD1(r, f, variable)


def Constraint_initialize_MBD1(f, variable):
    f.write("c" + " = 1 " + "\n")

    res = []
    for i in range (STATE_LENGTH):
        res.append(GetVariables(16,"BX0",STATE_LENGTH,variable)[i])
        res.append(GetVariables(16,"BX1",STATE_LENGTH,variable)[i])
    f.write(" + ".join(res) + " >= 1 " + "\n")

    for i in range (STATE_LENGTH):
        f.write(GetVariables(16,"BX0",STATE_LENGTH,variable)[i] + " = 0\n")
        f.write(GetVariables(16,"BX1",STATE_LENGTH,variable)[i] + " - " + GetVariables(16,"X",STATE_LENGTH,variable)[i] + " = 0\n")

def Constraint_sbox_MFD1(r, f, variable):
    P = [0, 33, 66, 99, 96, 1, 34, 67, 64, 97, 2, 35, 32, 65, 98, 3, 4, 37, 70, 103, 100, 5, 38, 71, 68, 101, 6, 39, 36, 69, 102, 7, 8, 41, 
    74, 107, 104, 9, 42, 75, 72, 105, 10, 43, 40, 73, 106, 11, 12, 45, 78, 111, 108, 13, 46, 79, 76, 109, 14, 47, 44, 77, 110, 15, 16, 49, 
    82, 115, 112, 17, 50, 83, 80, 113, 18, 51, 48, 81, 114, 19, 20, 53, 86, 119, 116, 21, 54, 87, 84, 117, 22, 55, 52, 85, 118, 23, 24, 57, 
    90, 123, 120, 25, 58, 91, 88, 121, 26, 59, 56, 89, 122, 27, 28, 61, 94, 127, 124, 29, 62, 95, 92, 125, 30, 63, 60, 93, 126, 31]

    M0 = [[0, 2, -1, -1, 0, 0, 0, 0, 2, 2, -7, -8, 3, 2, 3, 2, 0], 
    [-1, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 3, 0, -1, 0, -1, -1], 
    [0, 2, -1, -1, 0, 0, 0, 0, 2, 2, 3, 2, 3, 2, -7, -8, 0], 
    [0, 1, 0, 0, 0, 0, 0, 0, 1, 1, 1, 1, -3, -3, 1, 1, 0], 
    [0, 1, 0, 0, 0, 0, 0, 0, -3, -3, 1, 1, 1, 1, 1, 1, 0], 
    [-1, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 3, 0, -1, -1], 
    [-1, -1, 0, 1, 0, 0, 0, 0, 1, 4, 0, -1, 0, -1, 0, -1, 0], 
    [0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1], 
    [0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1], 
    [0, -1, 0, 0, 0, 1, 0, 0, 1, 0, 0, 0, 0, 0, 0, 1, 0], 
    [0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], 
    [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, -1], 
    [0, 0, 0, 1, 0, 0, 0, 0, 1, 0, -1, 0, 0, 0, 0, 0, 0], 
    [0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 1, 0, 0, 0, -1, 0, 0], 
    [0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0], 
    [0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0], 
    [0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1], 
    [0, 0, 0, 0, 0, 1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0], 
    [0, -1, 1, 1, 1, 0, 1, 0, 0, 0, -1, 0, -1, 0, -1, 0, -3], 
    [1, 1, 1, 1, 1, 0, 1, 0, -1, 0, 0, 1, 0, 1, 0, 1, 0], 
    [0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 1, 0, 0, 0, 0, 0, 0], 
    [0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 0], 
    [0, 0, 0, 0, 0, 0, 0, 1, -1, 0, 0, 0, 1, 0, 0, 0, 0], 
    [0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0], 
    [1, 0, 0, -1, 0, -1, 1, 1, -1, 0, 0, 0, 0, 0, 0, 0, -2], 
    [0, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0], 
    [0, 0, 0, 0, 0, 0, 0, -1, 1, 0, 0, 0, 0, 0, 0, 0, 0]]
    for i in range (int(STATE_LENGTH/4)):
        for t in range (len(M0)):
            res = []
            res.append(str(M0[t][0]) + " " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+3])
            res.append(str(M0[t][1]) + " " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+3])
            res.append(str(M0[t][2]) + " " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+2])
            res.append(str(M0[t][3]) + " " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+2])
            res.append(str(M0[t][4]) + " " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+1])
            res.append(str(M0[t][5]) + " " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+1])
            res.append(str(M0[t][6]) + " " + GetVariables(r,"BX0",STATE_LENGTH,variable)[4*i+0])
            res.append(str(M0[t][7]) + " " + GetVariables(r,"BX1",STATE_LENGTH,variable)[4*i+0])
            res.append(str(M0[t][8]) + " " + GetVariables(r+1,"BX0",STATE_LENGTH,variable)[P[4*i+3]])
            res.append(str(M0[t][9]) + " " + GetVariables(r+1,"BX1",STATE_LENGTH,variable)[P[4*i+3]])
            res.append(str(M0[t][10]) + " " + GetVariables(r+1,"BX0",STATE_LENGTH,variable)[P[4*i+2]])
            res.append(str(M0[t][11]) + " " + GetVariables(r+1,"BX1",STATE_LENGTH,variable)[P[4*i+2]])
            res.append(str(M0[t][12]) + " " + GetVariables(r+1,"BX0",STATE_LENGTH,variable)[P[4*i+1]])
            res.append(str(M0[t][13]) + " " + GetVariables(r+1,"BX1",STATE_LENGTH,variable)[P[4*i+1]])
            res.append(str(M0[t][14]) + " " + GetVariables(r+1,"BX0",STATE_LENGTH,variable)[P[4*i+0]])
            res.append(str(M0[t][15]) + " " + GetVariables(r+1,"BX1",STATE_LENGTH,variable)[P[4*i+0]])
            f.write(" + ".join(res) + " - " + str(M0[t][16]) + " c" + " >= 0 " + "\n")

    for i in range (STATE_LENGTH):
        f.write(GetVariables(r,"BY0",STATE_LENGTH,variable)[i] + " - " + GetVariables(r+1,"BX0",STATE_LENGTH,variable)[P[i]] + " = 0\n")
        f.write(GetVariables(r,"BY1",STATE_LENGTH,variable)[i] + " - " + GetVariables(r+1,"BX1",STATE_LENGTH,variable)[P[i]] + " = 0\n")

def Constraint_M_BD1(f, variable):
    Constraint_initialize_MBD1(f, variable)
    for r in range (16, 18):
        Constraint_sbox_MFD1(r, f, variable)

def Constraint_others(f, variable):
    M0 = [[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, -1, 3, 0],
[0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 1, 0, 0, 0],
[0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 1, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 1, 1, 0, 1, 0],
[0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 1, 0, 1, 0],
[1, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 1, 0, 0, 0, 0, 0, 1, 1, 1, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 0, 1, 1, 1, 0],
[0, 0, -1, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 1, 1, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 1, 0, 1, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 1, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 1, 0, 0, 0],
[0, -1, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 0, 1, -1, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 1, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 1, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 1, 1, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 1, 1, 0, 1, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 1, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 1, 0, 1, 1, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 1, 0, 1, 1, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 1, 1, 0],
[1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 0, 1, 1, 1, -1, 0, 0, 0, 0, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 1, 0, 0, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 1, 1, 0, 0, 0, 0, 0, 0, 1, -1, 2, 0],
[0, 0, 1, 1, 0, 0, 1, 1, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 1, 2, 0],
[0, 0, 1, 1, -1, 0, 0, 0, 0, 0, 1, 1, -1, 0, 0, 0, 0, 0, -1, 1, 2, 0],
[0, 0, -1, 0, 0, 0, 1, 1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, -1, 2, 0],
[0, 0, 0, -1, 0, 0, 1, 1, 0, 1, 1, 1, -1, 0, 0, 0, 0, 0, 1, 1, 1, 0],
[0, 0, -1, 0, 0, 1, 1, 1, 0, 0, 1, 1, -1, 0, 0, 0, 0, 0, 1, 1, 1, 0],
[-1, 0, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, -1, 0, 0, 0, 0, 0, -1, 1, 2, 0],
[0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, -1, 2, 0],
[1, 0, 0, 0, 1, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 0, 0, 1, 1, 0, 0, 0],
[0, 0, 1, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 1, 1, 0],
[-1, 0, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, 0, 0, -1, 0, 0, 0, 1, 1, 1, 0],
[1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 0, 1, 1, 0, 1, 0, -1, 0, 0, 0, 0],
[0, 0, 1, 1, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 1, 0, 0, -1, 1, 2, 0],
[-1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 1, 1, 0, 0, 0],
[0, 0, -1, 0, 0, 1, 1, 1, -1, 0, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, 1, 0],
[0, 0, 0, -1, 0, 0, 1, 1, -1, 0, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, 1, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, -1, 0, 0, 0, 0, 0, 1, 1, 1, 0],
[0, 0, 1, 1, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, 1, 0],
[0, 1, 1, 1, 0, 0, 1, 1, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 1, 1, 0],
[-1, 0, 0, 0, 0, 0, 1, 1, -1, 0, 0, 0, 0, 0, 1, 1, 0, 0, -1, 1, 2, 0],
[0, 1, 1, 0, 0, 0, 0, 0, 0, 1, 1, 0, 0, 0, 0, -1, 0, 0, 1, -1, 1, 0],
[-1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 1, 0, 1, 1, 1, 0, 0, 1, 1, 1, 0],
[-1, 0, 0, 0, 0, 1, 1, 1, 0, 0, 1, 1, 0, 0, 0, -1, 0, 0, 1, 1, 1, 0],
[1, 0, 0, 0, 1, 0, 1, 0, 1, 0, 1, 0, 0, 0, 1, 1, 0, 1, 0, -1, 0, 0],
[0, 0, 0, 1, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 0],
[0, 1, 0, 0, -1, 0, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, -1, -1, 2, 0],
[0, 1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, -1, 0, 0, 0, 0, 0, -1, -1, 2, 0],
[0, 1, 0, 1, 0, 0, -1, 0, 0, 1, 1, 1, 0, 1, 0, 1, 0, 0, 0, -1, 1, 0],
[0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0, -1, 0, 0, 0, 0, 1, 1, 1, 0],
[1, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 1, 1, 0, 0, 0],
[0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0, 0, 1, 1, 1, 0],
[-1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, 0, 0, -1, 1, 2, 0],
[0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, -1, 0, 0, 0],
[-1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 1, 1, 0],
[0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, 0, 0, 0, 0, -1, 1, 0, 0, 0],
[0, 0, 0, 1, 0, 0, 1, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 1, 1, 1, 0],
[0, 1, 1, 0, 0, 1, 0, 1, 0, -1, 0, 0, 0, 1, 1, 1, 0, 0, 1, -1, 1, 0],
[0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0, 0, 1, 1, 1, 0],
[-1, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0, 0, -1, 0, 0, 0, 0, 1, 1, 1, 0],
[1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 1, 1, 1, 0, 0, 1, 0, 0, 1, -1, 0, 0],
[0, 0, 0, 0, 0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, -1, 2, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 0, 0, 0, 0, 0, -1, -1, -1, 2, 0],
[0, 1, 1, 1, 0, 1, 1, 1, 0, 0, 0, 0, 0, 1, 1, 1, 0, -1, 1, 0, 0, 0],
[0, 1, 1, 1, 0, 0, 0, 0, 0, 1, 1, 1, 0, 1, 1, 1, 0, -1, 1, 0, 0, 0],
[0, 0, 1, 1, 0, 1, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 0, 0, 1, -1, 1, 0],
[0, 1, 0, 1, 0, 1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 1, 0, 0, 1, -1, 1, 0],
[1, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 0, -1, -1, 2, 0],
[1, 0, 0, 0, 1, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 0, 0, 0, -1, -1, 2, 0],
[1, 1, 1, 1, 1, 0, 0, 1, 1, 0, 0, 0, 1, 0, 0, 0, 0, -1, 0, -1, 1, 0],
[0, 0, 0, 0, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 0, 1, 0, 0, 1, -1, 0, 0],
[0, 1, 1, 1, 0, 1, 1, 0, 0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 1, -1, 0, 0],
[0, 1, 0, 1, 0, 0, 0, 0, 0, 1, 1, 1, 0, 1, 1, 1, 0, 0, 1, -1, 0, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 0, -1, -1, -1, 2, 0],
[-1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 0, -1, -1, 2, 0],
[0, 0, 0, 0, 0, 1, 1, 1, 0, 1, 1, 1, 0, 1, 1, 1, 0, -1, 1, 0, 0, 0],
[0, 1, 1, 1, 0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, -1, 0, 2, 0],
[0, 1, 1, 1, 0, 0, 0, 0, 0, 1, 1, 1, 0, 0, 0, 0, -1, -1, -1, 0, 2, 0],
[0, 1, 1, 1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, -1, -1, -1, 0, 2, 0],
[0, 0, 0, 0, 0, 1, 1, 1, 0, 1, 1, 1, 0, 0, 0, 0, -1, -1, -1, 0, 2, 0],
[0, 0, 0, 0, 0, 1, 1, 1, 0, 0, 0, 0, 0, 1, 1, 1, -1, -1, -1, 0, 2, 0],
[1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 1, 0, 0, 0, 0, 1, -1, 1, 0, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 1, 0, 1, 1, 1, -1, -1, -1, 0, 2, 0],
[0, 1, 1, 0, 0, 1, 1, 1, 0, 1, 0, 1, 0, 1, 0, 0, 0, 0, 0, -1, 0, 0],
[0, 1, 0, 0, 0, 1, 1, 0, 0, 1, 1, 0, 0, 1, 1, 1, 0, 0, 0, -1, 0, 0],
[0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 0],
[-1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[-1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[-1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 2, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 2, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 2, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 2, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 2, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 2, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 2, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 2, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 3, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 2, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0],
[0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, -1, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, -1, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 1, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 1, 0, 0, 0, 0, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 1, 0, 0, 0, 0, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 1, 0, 0, 0, 0, 0],
[0, 0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0],
[0, 0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0],
[0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0],
[0, 0, 0, 0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0],
[0, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0],
[0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0],
[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, -1, -1, 0, 0, 0, 0, 1, 0],
[0, 0, 0, 0, 0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, 0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, -1, 0, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],
[0, -1, -1, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 1, 0],]

    for i in range (8):
        for t in range (len(M0)):
            res = []
            res.append(str(M0[t][0]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+15])
            res.append(str(M0[t][1]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+14])
            res.append(str(M0[t][2]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+13])
            res.append(str(M0[t][3]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+12])
            res.append(str(M0[t][4]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+11])
            res.append(str(M0[t][5]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+10])
            res.append(str(M0[t][6]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+9])
            res.append(str(M0[t][7]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+8])
            res.append(str(M0[t][8]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+7])
            res.append(str(M0[t][9]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+6])
            res.append(str(M0[t][10]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+5])
            res.append(str(M0[t][11]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+4])
            res.append(str(M0[t][12]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+3])
            res.append(str(M0[t][13]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+2])
            res.append(str(M0[t][14]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+1])
            res.append(str(M0[t][15]) + " " + GetVariables(16,"X",STATE_LENGTH,variable)[16*i+0])
            res.append(str(M0[t][16]) + " " + GetVariables(0,"M",32,variable)[4*i+3])
            res.append(str(M0[t][17]) + " " + GetVariables(0,"M",32,variable)[4*i+2])
            res.append(str(M0[t][18]) + " " + GetVariables(0,"M",32,variable)[4*i+1])
            res.append(str(M0[t][19]) + " " + GetVariables(0,"M",32,variable)[4*i+0])
            f.write(" + ".join(res) + " + " + str(M0[t][20]) + " c" + " >= 0 " + "\n")

def Constraint(f, variable):
    Constraint_FD0(f, variable)
    Constraint_M_BD0(f, variable)
    Constraint_FD1(f, variable)
    Constraint_M_BD1(f, variable)
    Constraint_others(f, variable)

def ObjectiveFunction(f, variable):
    res = []
    for i in range (STATE_LENGTH):
        res.append(GetVariables(19,"KK",STATE_LENGTH,variable)[i])
    for i in range (8):
        res.append(" 8 " + GetVariables(0,"M",32,variable)[4*i+3])
        res.append(" 4 " + GetVariables(0,"M",32,variable)[4*i+2])
        res.append(" 2 " + GetVariables(0,"M",32,variable)[4*i+1])
        res.append(" 1 " + GetVariables(0,"M",32,variable)[4*i+0])
    f.write(" + ".join(res) + "\n")

def VariablesType(f):
    f.write("\n".join(variable) + "\n")

def CreateModel(lpFileName, variable):
    f = open(lpFileName, "w")
    f.write("Minimum\n")
    ObjectiveFunction(f, variable)
    f.write("Subject To\n")
    Constraint(f, variable)
    f.write("Binaries\n")
    VariablesType(f)
    f.write("End\n")
    f.close()

def SolveModel(lpFileName, solFileName, aa_data, index):
    model = read(lpFileName)
    model.setParam('OutputFlag', 1) 
    model.Params.PoolSearchMode = 2
    model.Params.PoolSolutions = 1
    model.optimize()


    with open("All_MILP_Results.txt", "a", encoding="utf-8") as txt_file:
        txt_file.write(f"################### Object {index} ###################\n")

        txt_file.write(f"Input Diff: {hex(aa_data[0])}\n")
        txt_file.write(f"Output Mask: {hex(aa_data[1])}\n")
        txt_file.write(f"Weight: {aa_data[2]}, Prob: {aa_data[3]}\n")

        if model.Status == GRB.OPTIMAL:
 
            current_sol = f"BAKSHEESH_{index}.sol"
            model.write(current_sol)
            
            obj_val = model.ObjVal
            txt_file.write(f"Status: OPTIMAL\n")
            txt_file.write(f"MILP Objective Value (Active Sboxes): {obj_val}\n")
            
            SaveXToExcel(model, f"Result_X_{index}.xlsx")
            SaveYToExcel(model, f"Result_Y_{index}.xlsx")
            SaveBXToExcel(model, f"Result_BX_{index}.xlsx")
            SaveBYToExcel(model, f"Result_BY_{index}.xlsx")
            SaveXXToExcel(model, f"Result_XX_{index}.xlsx")
            SaveKKToExcel(model, f"Result_KK_{index}.xlsx")
            
            txt_file.write(f"Excel matrices and .sol file generated successfully.\n")
        else:
            txt_file.write(f"Status: NOT OPTIMAL (Gurobi Status Code: {model.Status})\n")
        
        txt_file.write("\n" + "="*50 + "\n\n")

if __name__ == '__main__':
    AA = [[0x20000000100, 0x4040000000000000000, 2,2, 48.0, 114.0242],
        [0x100, 0x9400000000000000000000000000000, 1,3, 48.0, 114.0995],
        [0x4, 0x940000000000000, 1,3, 48.0, 114.1790],
        [0x10000000000000000, 0x9400000000000000000, 1,3, 48.0, 114.1790],
        [0x1000000000000000000000000000, 0x4009000000000000000000000000, 1,3, 48.0, 114.1790],
        [0x20000000100, 0x40400000000, 2,2, 48.0, 114.2065],
        [0x400000000000000000000000000, 0x4009, 1,3, 48.0, 114.2065],
        [0x1000000000000000000, 0x940, 1,3, 48.0, 114.3523],
        [0x10000000000000000000, 0x940, 1,3, 48.0, 114.3523],
        [0x1000000000000000000000000000, 0x4009, 1,3, 48.0, 114.4150],
        [0x1000, 0x9400000, 1,3, 48.0, 114.4806],
        [0x100000000000000000000000000, 0x4009000000000000000000000000, 1,3, 48.0, 114.6215],
        [0x400000000000000000000000000, 0x4009000000000000000000000000, 1,3, 48.0, 114.6215],
        [0x200000001000000000000000000, 0x40400000000, 2,2, 48.0, 114.6590],
        [0x10000000000000000000000000, 0x400900000000, 1,3, 48.0, 114.7776],
        [0x4000000000000000100, 0x404000000000000, 2,2, 48.0, 114.8625],
        [0x1, 0x940000000000000, 1,3, 48.0,  114.9527],
        [0x100000000, 0x40090000000000000000, 1,3, 48.0, 114.9527]]
    
    with open("All_MILP_Results.txt", "a", encoding="utf-8") as f:
        f.write("=== MILP Automated Search Experiment Log ===\n\n")

    STATE_LENGTH = 128

    for aa in range(len(AA)):
        variable = set()
        lpFileName = "BAKSHEESH.lp"
        solFileName = f"BAKSHEESH_{aa}.sol" 

        CreateModel(lpFileName, variable) 
        SolveModel(lpFileName, solFileName, AA[aa], aa)